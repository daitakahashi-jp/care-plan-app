import json, re, io, os
from datetime import datetime
from http.server import BaseHTTPRequestHandler
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "..", "template.xlsx")

def fill_plan(data: dict, wb) -> None:
    ws = wb.active
    ui       = data.get("userInfo", {})
    goals    = data.get("goals", {})
    services = data.get("weeklyServices", {})
    created  = data.get("createdAt", "")

    def w(cell, value, size=8):
        c = ws[cell]; c.value = value
        c.font = Font(size=size, name="メイリオ")

    def wa(cell, value, size=8):
        c = ws[cell]; c.value = value
        c.font = Font(size=size, name="メイリオ")
        c.alignment = Alignment(wrap_text=True, vertical="top")

    try:
        dt = datetime.fromisoformat(created[:10]) if created else datetime.today()
    except:
        dt = datetime.today()

    w("H1", dt.year, 9);  w("K1", dt.month, 9); w("M1", dt.day, 9)
    w("H2", ui.get("svcresp", ""), 9)
    w("E3", ui.get("name", ""), 10)
    gender = ui.get("gender", "")
    w("N3", "■男　　□ 女" if gender == "男" else "□男　　■ 女", 9)
    birth = ui.get("birth", "")
    w("V3", "□明治　□大正　■昭和" if "昭和" in birth else "□明治　□大正　□昭和", 8)
    m = re.search(r'(\d+)年\s*(\d+)月\s*(\d+)日', birth)
    if m:
        w("AB3", int(m.group(1))); w("AD3", int(m.group(2))); w("AF3", int(m.group(3)))
    w("AG3", ui.get("addr", ""), 8)
    w("E4",  ui.get("emg1_name", ""), 9)
    w("L4",  "続柄：" + ui.get("emg1_rel", ""), 8)
    w("S4",  ui.get("emg1_tel", ""), 8)
    w("AD4", ui.get("emg2_name", ""), 9)
    w("AK4", "続柄：" + ui.get("emg2_rel", ""), 8)
    w("AR4", ui.get("emg2_tel", ""), 8)
    w("E5",  ui.get("office", ""), 9)
    w("O5",  ui.get("office_tel", ""), 8)
    w("V5",  ui.get("office_addr", ""), 8)
    w("AM5", ui.get("manager", ""), 9)
    w("E6",  ui.get("svcresp", ""), 9)
    w("T6",  ui.get("helper", ""), 9)
    w("L7",  ui.get("cm_name","") + "　" + ui.get("cm_office",""), 8)
    w("Z7",  ui.get("cm_tel", ""), 8)
    w("AH7", ui.get("doctor", ""), 8)
    w("AR7", ui.get("doctor_tel", ""), 8)

    period = ui.get("period", "")
    pm = re.findall(r'(\d{4})年(\d{1,2})月(\d{1,2})日', period)
    if len(pm) >= 2:
        w("H9",int(pm[0][0])); w("J9",int(pm[0][1])); w("L9",int(pm[0][2]))
        w("O9",int(pm[1][0])); w("Q9",int(pm[1][1])); w("S9",int(pm[1][2]))

    day_start_col = {"月":3,"火":8,"水":13,"木":18,"金":23,"土":28,"日":33}
    time_rows = {6:12,7:14,8:16,9:18,10:20,11:22,12:24,13:26,14:28,
                 15:30,16:32,17:34,18:36,19:38,20:40,21:42,22:44,23:46}
    body_row = time_rows[9]
    life_row = time_rows[14]

    def get_parent(addr):
        for mr in ws.merged_cells.ranges:
            cells = [f"{get_column_letter(cc)}{rr}"
                     for rr in range(mr.min_row, mr.max_row+1)
                     for cc in range(mr.min_col, mr.max_col+1)]
            if addr in cells:
                return f"{get_column_letter(mr.min_col)}{mr.min_row}"
        return addr

    for svc_name, svc_info in services.items():
        svc_type = svc_info.get("type","")
        svc_days = svc_info.get("days",[])
        row  = body_row if svc_type == "身体介護" else life_row
        icon = "▶" if svc_type == "身体介護" else "◆"
        label = f"{icon}{svc_name}"
        for day in svc_days:
            sc = day_start_col.get(day)
            if sc is None: continue
            addr = get_parent(f"{get_column_letter(sc)}{row}")
            existing = ws[addr].value
            ws[addr].value = (str(existing) + "\n" + label) if existing else label
            ws[addr].font = Font(size=7, name="メイリオ")
            ws[addr].alignment = Alignment(wrap_text=True, vertical="top")
        if svc_type == "身体介護":
            body_row = min(body_row + 2, life_row - 2)
        else:
            life_row = min(life_row + 2, 46)

    notes_list = goals.get("notes", [])
    notes_str = "\n".join(f"・{n}" for n in (notes_list if isinstance(notes_list, list) else [str(notes_list)]))
    goal_text = (
        f"【長期目標】\n{goals.get('long','')}\n\n"
        f"【短期目標①】\n{goals.get('short1','')}\n\n"
        f"【短期目標②】\n{goals.get('short2','')}\n\n"
        f"【留意事項】\n{notes_str}"
    )
    wa("AM10", goal_text, 8)
    ws.row_dimensions[10].height = 200
    for r in range(11, 48):
        if ws.row_dimensions[r].height is None or ws.row_dimensions[r].height < 25:
            ws.row_dimensions[r].height = 25


class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200); self._cors(); self.end_headers()

    def do_POST(self):
        length = int(self.headers.get("Content-Length", 0))
        try:
            data = json.loads(self.rfile.read(length))
        except Exception as e:
            self._json(400, {"error": str(e)}); return

        try:
            wb = load_workbook(TEMPLATE_PATH)
            fill_plan(data, wb)
            out = io.BytesIO()
            wb.save(out)
            xlsx = out.getvalue()
        except Exception as e:
            self._json(500, {"error": str(e)}); return

        self.send_response(200); self._cors()
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", "attachment; filename*=UTF-8''%E8%A8%AA%E5%95%8F%E4%BB%8B%E8%AD%B7%E8%A8%88%E7%94%BB%E6%9B%B8.xlsx")
        self.send_header("Content-Length", str(len(xlsx)))
        self.end_headers(); self.wfile.write(xlsx)

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def _json(self, status, obj):
        body = json.dumps(obj).encode()
        self.send_response(status); self._cors()
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers(); self.wfile.write(body)

    def log_message(self, *args): pass
